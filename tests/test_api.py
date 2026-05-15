"""
API endpoint tests for YusBuild.
Tests all CRUD operations and custom actions.
"""

import json

import pytest
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from rest_framework import status
from rest_framework.test import APIClient

from apps.piles.models import (
    Pile,
    PileCalculationHistory,
)
from apps.projects.models import Project, ProjectMembership


@pytest.fixture
def authenticated_user(db):
    user = get_user_model().objects.create_user(
        username="test-user",
        password="test-password",
    )
    engineer_group, _ = Group.objects.get_or_create(name="engineer")
    user.groups.add(engineer_group)
    return user


@pytest.fixture
def api_client(authenticated_user):
    client = APIClient()
    client.force_authenticate(user=authenticated_user)
    return client


@pytest.fixture
def project(db, authenticated_user):
    project = Project.objects.create(
        name="Refinery Extension Test Pile",
        location="Crude Distillation Unit",
        client="Engineers India Limited",
        description="Residential development at Lekki, Lagos",
        status="ACTIVE",
        created_by="Engr. Yusuf",
    )
    ProjectMembership.objects.create(
        project=project,
        user=authenticated_user,
        role=ProjectMembership.ROLE_ENGINEER,
    )
    return project


class TestHealthEndpoint:
    """Tests for health check endpoint."""

    def test_health_check(self, api_client):
        """Health endpoint should return 200 with status ok."""
        response = api_client.get("/health/")
        assert response.status_code == status.HTTP_200_OK
        data = json.loads(response.content)
        assert data["status"] == "ok"
        assert data["service"] == "yusbuild-api"

    def test_health_check_includes_request_id(self, api_client):
        """Responses should include a request ID for tracing."""
        response = api_client.get("/health/")

        assert response.status_code == status.HTTP_200_OK
        assert response["X-Request-ID"]

    def test_health_check_preserves_inbound_request_id(self, api_client):
        """Request ID middleware should preserve upstream request IDs."""
        response = api_client.get("/health/", HTTP_X_REQUEST_ID="req-test-123")

        assert response["X-Request-ID"] == "req-test-123"

    def test_readiness_check(self, api_client):
        """Readiness endpoint should verify runtime dependencies."""
        response = api_client.get("/readiness/")

        assert response.status_code == status.HTTP_200_OK


class TestPileBOQExport:
        @pytest.mark.django_db
        def test_boq_export_csv_empty(self, api_client):
            url = "/api/v1/piles/boq-export-csv/"
            response = api_client.get(url)
            assert response.status_code == 200
            content = response.content.decode()
            assert "No data" in content

        @pytest.mark.django_db
        def test_boq_export_xlsx_empty(self, api_client):
            url = "/api/v1/piles/boq-export-xlsx/"
            response = api_client.get(url)
            assert response.status_code == 200
            import io
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(response.content))
            ws = wb.active
            assert ws["A1"].value == "No data"
    class TestPileCSVImport:
    @pytest.mark.django_db
    def test_import_csv_no_file(self, api_client):
        response = api_client.post("/api/v1/piles/import-csv/", {}, format="multipart")
        assert response.status_code == 400
        assert "No file uploaded" in response.json()["detail"]

    @pytest.mark.django_db
    def test_import_csv_empty_file(self, api_client):
        import io
        file = io.BytesIO(b"")
        file.name = "empty.csv"
        response = api_client.post("/api/v1/piles/import-csv/", {"file": file}, format="multipart")
        # Should not crash, but will have no rows
        assert response.status_code == 200 or response.status_code == 400

            class TestPileBulkCreate:
                    @pytest.mark.django_db
                    def test_bulk_create_not_list(self, api_client):
                        payload = {"pile_no": "P-999"}
                        response = api_client.post("/api/v1/piles/bulk-create/", payload, format="json")
                        assert response.status_code == 400
                        assert "Expected a list" in response.json()["detail"]

                    @pytest.mark.django_db
                    def test_bulk_create_empty_list(self, api_client):
                        response = api_client.post("/api/v1/piles/bulk-create/", [], format="json")
                        assert response.status_code == 200
                        data = response.json()
                        assert data["created"] == []
                        assert data["errors"] == []
                """Test bulk pile creation endpoint."""

                @pytest.mark.django_db
                def test_bulk_create_success(self, api_client, project):
                    payload = [
                        {
                            "pile_no": "P-500",
                            "pile_type": "BORED",
                            "diameter_mm": 900,
                            "design_length_m": 22.0,
                            "actual_length_m": 21.5,
                            "project": project.id,
                        },
                        {
                            "pile_no": "P-501",
                            "pile_type": "BORED",
                            "diameter_mm": 1000,
                            "design_length_m": 25.0,
                            "actual_length_m": 24.5,
                            "project": project.id,
                        },
                    ]
                    response = api_client.post("/api/v1/piles/bulk-create/", payload, format="json")
                    assert response.status_code == 200
                    data = response.json()
                    assert "created" in data
                    assert len(data["created"]) == 2
                    assert not data["errors"]

                @pytest.mark.django_db
                def test_bulk_create_with_errors(self, api_client, project):
                    payload = [
                        {
                            "pile_no": "P-600",
                            "pile_type": "BORED",
                            "diameter_mm": 900,
                            "design_length_m": 22.0,
                            "actual_length_m": 21.5,
                            "project": project.id,
                        },
                        {
                            "pile_no": "P-601",
                            "pile_type": "BORED",
                            "diameter_mm": "INVALID",
                            "design_length_m": 25.0,
                            "actual_length_m": 24.5,
                            "project": project.id,
                        },
                    ]
                    response = api_client.post("/api/v1/piles/bulk-create/", payload, format="json")
                    assert response.status_code == 400
                    data = response.json()
                    assert "errors" in data
                    assert len(data["errors"]) == 1
                    assert data["errors"][0]["row"] == 2
                    assert "diameter_mm" in str(data["errors"][0]["errors"])  # Should mention the invalid field
            @pytest.mark.django_db
            def test_import_csv_dry_run_success(self, api_client, project):
                import io
                csv_content = (
                    "pile_no,pile_type,diameter_mm,design_length_m,actual_length_m,project\n"
                    "P-300,BORED,800,20.0,19.5,{project_id}\n"
                    "P-301,BORED,750,18.0,17.5,{project_id}\n"
                ).format(project_id=project.id)
                file = io.BytesIO(csv_content.encode("utf-8"))
                file.name = "piles.csv"
                response = api_client.post(
                    "/api/v1/piles/import-csv/?dry_run=1",
                    {"file": file},
                    format="multipart",
                )
                assert response.status_code == 200
                data = response.json()
                assert data["dry_run"] is True
                assert len(data["created"]) == 2
                assert all(r["status"] == "valid" for r in data["created"])
                assert not data["errors"]

            @pytest.mark.django_db
            def test_import_csv_dry_run_with_errors(self, api_client, project):
                import io
                csv_content = (
                    "pile_no,pile_type,diameter_mm,design_length_m,actual_length_m,project\n"
                    "P-400,BORED,800,20.0,19.5,{project_id}\n"
                    "P-401,BORED,INVALID,18.0,17.5,{project_id}\n"
                ).format(project_id=project.id)
                file = io.BytesIO(csv_content.encode("utf-8"))
                file.name = "piles.csv"
                response = api_client.post(
                    "/api/v1/piles/import-csv/?dry_run=true",
                    {"file": file},
                    format="multipart",
                )
                assert response.status_code == 400
                data = response.json()
                assert data["dry_run"] is True
                assert len(data["errors"]) == 1
                assert data["errors"][0]["row"] == 3
        """Test pile schedule CSV import endpoint."""

        @pytest.mark.django_db
        def test_import_csv_success(self, api_client, project):
            import io
            csv_content = (
                "pile_no,pile_type,diameter_mm,design_length_m,actual_length_m,project\n"
                "P-100,BORED,800,20.0,19.5,{project_id}\n"
                "P-101,BORED,750,18.0,17.5,{project_id}\n"
            ).format(project_id=project.id)
            file = io.BytesIO(csv_content.encode("utf-8"))
            file.name = "piles.csv"
            response = api_client.post(
                "/api/v1/piles/import-csv/",
                {"file": file},
                format="multipart",
            )
            assert response.status_code == 200
            data = response.json()
            assert "created" in data
            assert len(data["created"]) == 2
            assert not data["errors"]

        @pytest.mark.django_db
        def test_import_csv_row_errors(self, api_client, project):
            import io
            csv_content = (
                "pile_no,pile_type,diameter_mm,design_length_m,actual_length_m,project\n"
                "P-200,BORED,800,20.0,19.5,{project_id}\n"
                "P-201,BORED,INVALID,18.0,17.5,{project_id}\n"
            ).format(project_id=project.id)
            file = io.BytesIO(csv_content.encode("utf-8"))
            file.name = "piles.csv"
            response = api_client.post(
                "/api/v1/piles/import-csv/",
                {"file": file},
                format="multipart",
            )
            assert response.status_code == 400
            data = response.json()
            assert "errors" in data
            assert len(data["errors"]) == 1
            assert data["errors"][0]["row"] == 3
            assert "diameter_mm" in str(data["errors"][0]["errors"])  # Should mention the invalid field
    """Test BOQ CSV export endpoint."""

    @pytest.mark.django_db
    def test_boq_export_csv(self, api_client, project):
        # Create a pile for export
        Pile.objects.create(
            pile_no="P-001",
            project=project,
            pile_type="BORED",
            diameter_mm=600,
            design_length_m=12.5,
            actual_length_m=12.0,
        )
        url = "/api/v1/piles/boq-export-csv/"
        response = api_client.get(url)
        assert response.status_code == 200
        assert response["Content-Type"].startswith("text/csv")
        content = response.content.decode()
        assert "pile_no" in content
        assert "P-001" in content

    @pytest.mark.django_db
    def test_boq_export_xlsx(self, api_client, project):
        # Create a pile for export
        Pile.objects.create(
            pile_no="P-002",
            project=project,
            pile_type="BORED",
            diameter_mm=750,
            design_length_m=15.0,
            actual_length_m=14.5,
        )
        url = "/api/v1/piles/boq-export-xlsx/"
        response = api_client.get(url)
        assert response.status_code == 200
        assert response["Content-Type"].startswith("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        assert response["Content-Disposition"].endswith("boq_export.xlsx"")
        # Check that the response is a valid Excel file by loading it
        import io
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        assert "pile_no" in headers
        found = False
        for row in ws.iter_rows(min_row=2, values_only=True):
            if "P-002" in row:
                found = True
        assert found
        data = json.loads(response.content)
        assert data["status"] == "ready"
        assert data["checks"]["database"] == "ok"
        assert data["checks"]["migrations"] == "ok"


class TestProjectEndpoints:
    def test_boq_csv_export(self, api_client, project):
        """GET /api/v1/projects/{id}/boq-csv/ should return a valid CSV export."""
        # Create a pile for the project
        pile = Pile.objects.create(
            project=project,
            pile_no="P-001",
            pile_type="TYPE_II",
            diameter_mm=500,
            design_length_m=20.0,
            actual_length_m=21.2,
        )
        # Ensure calculation exists
        pile.calculation = pile.calculation or None
        response = api_client.get(f"/api/v1/projects/{project.id}/boq-csv/")
        assert response.status_code == 200
        assert response["Content-Type"].startswith("text/csv")
        assert f'attachment; filename="boq_{project.id}.csv"' in response["Content-Disposition"]
        content = response.content.decode()
        # Check CSV headers
        assert "Pile No,Pile Type,Diameter (mm),Design Length (m),Actual Length (m),Steel (kg),Steel (tons),Concrete (m3),Main Bars (kg),Helix (kg),Stiffeners (kg)" in content
        # Check pile data row
        assert "P-001" in content
    """Tests for Project CRUD endpoints."""

    def test_list_projects(self, api_client, project):
        """GET /api/v1/projects/ should return project list."""
        response = api_client.get("/api/v1/projects/")
        assert response.status_code == status.HTTP_200_OK
        assert response.data["count"] >= 1

    def test_create_project(self, api_client):
        """POST /api/v1/projects/ should create project."""
        data = {
            "name": "New Bridge Project",
            "location": "Lekki, Lagos",
            "client": "Lekki Development Corp",
            "description": "Bridge foundation works",
            "status": "ACTIVE",
            "created_by": "Engr. Yusuf",
        }
        response = api_client.post("/api/v1/projects/", data, format="json")
        assert response.status_code == status.HTTP_201_CREATED
        assert response.data["name"] == "New Bridge Project"

    def test_get_project(self, api_client, project):
        """GET /api/v1/projects/{id}/ should return project."""
        response = api_client.get(f"/api/v1/projects/{project.id}/")
        assert response.status_code == status.HTTP_200_OK
        assert response.data["name"] == project.name

    def test_update_project(self, api_client, project):
        """PUT /api/v1/projects/{id}/ should update project."""
        data = {
            "name": "Updated Name",
            "location": project.location,
            "client": project.client,
            "status": project.status,
        }
        response = api_client.put(
            f"/api/v1/projects/{project.id}/", data, format="json"
        )
        assert response.status_code == status.HTTP_200_OK
        assert response.data["name"] == "Updated Name"

    def test_delete_project(self, api_client, project):
        """DELETE /api/v1/projects/{id}/ should delete project."""
        response = api_client.delete(f"/api/v1/projects/{project.id}/")
        assert response.status_code == status.HTTP_204_NO_CONTENT
        assert not Project.objects.filter(id=project.id).exists()


class TestPileEndpoints:
    """Tests for Pile CRUD endpoints."""

    def test_list_piles(self, api_client, pile_type_ii):
        """GET /api/v1/piles/ should return pile list."""
        response = api_client.get("/api/v1/piles/")
        assert response.status_code == status.HTTP_200_OK
        assert response.data["count"] >= 1

    def test_create_pile(self, api_client, project, type_ii_config):
        """POST /api/v1/piles/ should create pile with auto-calculation."""
        data = {
            "pile_no": "P-NEW-001",
            "pile_type": "TYPE_II",
            "project": project.id,
            "diameter_mm": 500,
            "design_length_m": 20.0,
            "actual_length_m": 21.2,
            "piling_method": "Driven Cast In-Situ",
            "concrete_grade": "C35/40",
        }
        response = api_client.post("/api/v1/piles/", data, format="json")
        assert response.status_code == status.HTTP_201_CREATED
        assert response.data["pile_no"] == "P-NEW-001"
        assert response.data["calculation_result"] is not None
        assert response.data["calculation_result"]["steel"]["total_kg"] > 0
        pile = Pile.objects.get(pile_no="P-NEW-001")
        assert pile.calculation_history.count() == 1
        assert pile.calculation_history.first().trigger == "create"

    def test_get_pile(self, api_client, pile_type_ii):
        """GET /api/v1/piles/{id}/ should return pile with calculation."""
        response = api_client.get(f"/api/v1/piles/{pile_type_ii.id}/")
        assert response.status_code == status.HTTP_200_OK
        assert response.data["pile_no"] == pile_type_ii.pile_no
        assert "calculation" in response.data

    def test_update_pile(self, api_client, pile_type_ii):
        """PATCH /api/v1/piles/{id}/ should update and recalculate."""
        data = {"actual_length_m": 22.0, "notes": "Updated depth after drilling"}
        response = api_client.patch(
            f"/api/v1/piles/{pile_type_ii.id}/", data, format="json"
        )
        assert response.status_code == status.HTTP_200_OK
        assert response.data["actual_length_m"] == 22.0
        assert pile_type_ii.calculation_history.count() == 1
        assert pile_type_ii.calculation_history.first().trigger == "update"

    def test_delete_pile(self, api_client, pile_type_ii):
        """DELETE /api/v1/piles/{id}/ should delete pile."""
        pile_id = pile_type_ii.id
        response = api_client.delete(f"/api/v1/piles/{pile_type_ii.id}/")
        assert response.status_code == status.HTTP_204_NO_CONTENT
        assert not Pile.objects.filter(id=pile_id).exists()


class TestPileCustomActions:
    """Tests for pile custom actions."""

    def test_recalculate(self, api_client, pile_type_ii):
        """POST /api/v1/piles/{id}/recalculate/ should force recalculation."""
        response = api_client.post(
            f"/api/v1/piles/{pile_type_ii.id}/recalculate/",
            {"reason": "QA recalculation"},
            format="json",
        )
        assert response.status_code == status.HTTP_200_OK
        assert "history_id" in response.data
        assert "result" in response.data
        assert response.data["result"]["steel"]["total_kg"] > 0
        history = PileCalculationHistory.objects.get(id=response.data["history_id"])
        assert history.trigger == "recalculate"
        assert history.reason == "QA recalculation"
        assert history.input_snapshot["pile_no"] == pile_type_ii.pile_no
        assert history.result_snapshot["steel"]["total_kg"] > 0

    def test_breakdown(self, api_client, pile_type_ii):
        """GET /api/v1/piles/{id}/breakdown/ should return full breakdown."""
        response = api_client.get(f"/api/v1/piles/{pile_type_ii.id}/breakdown/")
        assert response.status_code == status.HTTP_200_OK
        assert "steel" in response.data
        assert "concrete" in response.data
        assert "main_bars" in response.data["steel"]
        assert "helix" in response.data["steel"]
        assert "stiffeners" in response.data["steel"]

    def test_calculation_history(self, api_client, pile_type_ii):
        """GET /api/v1/piles/{id}/calculation-history/ returns audit history."""
        api_client.post(f"/api/v1/piles/{pile_type_ii.id}/recalculate/")

        response = api_client.get(
            f"/api/v1/piles/{pile_type_ii.id}/calculation-history/"
        )

        assert response.status_code == status.HTTP_200_OK
        assert response.data["count"] == 1
        item = response.data["results"][0]
        assert item["trigger"] == "recalculate"
        assert item["input_snapshot"]["pile_no"] == pile_type_ii.pile_no
        assert item["config_snapshot"]["pile_type"] == pile_type_ii.pile_type


class TestBOQEndpoint:
    """Tests for BOQ generation endpoint."""

    def test_boq_empty_project(self, api_client, project):
        """BOQ for project with no piles should return empty."""
        response = api_client.get(f"/api/v1/projects/{project.id}/boq/")
        assert response.status_code == status.HTTP_200_OK
        assert response.data["grand_totals"]["total_piles"] == 0

    def test_boq_with_piles(self, api_client, pile_type_ii):
        """BOQ should aggregate pile data correctly."""
        project_id = pile_type_ii.project.id
        response = api_client.get(f"/api/v1/projects/{project_id}/boq/")
        assert response.status_code == status.HTTP_200_OK
        assert response.data["grand_totals"]["total_piles"] >= 1
        assert response.data["grand_totals"]["total_steel_kg"] > 0
        assert "summary_by_type" in response.data
        assert "steel_distribution" in response.data

    def test_boq_summary_by_type(self, api_client, pile_type_ii, pile_type_iii):
        """BOQ should group by pile type."""
        project_id = pile_type_ii.project.id
        response = api_client.get(f"/api/v1/projects/{project_id}/boq/")
        assert response.status_code == status.HTTP_200_OK

        summary = response.data["summary_by_type"]
        assert len(summary) >= 2  # At least TYPE_II and TYPE_III

    def test_boq_steel_distribution(self, api_client, pile_type_ii):
        """BOQ should include steel distribution percentages."""
        project_id = pile_type_ii.project.id
        response = api_client.get(f"/api/v1/projects/{project_id}/boq/")
        assert response.status_code == status.HTTP_200_OK

        dist = response.data["steel_distribution"]
        assert "main_bars" in dist
        assert "helix" in dist
        assert "stiffeners" in dist
        assert dist["main_bars"]["percentage"] > 0


class TestPileTypeConfigEndpoints:
    """Tests for pile type configuration endpoints."""

    def test_list_configs(self, api_client, type_ii_config):
        """GET /api/v1/piles/configs/ should return configurations."""
        response = api_client.get("/api/v1/piles/configs/")
        assert response.status_code == status.HTTP_200_OK
        assert response.data["count"] >= 1

    def test_get_config(self, api_client, type_ii_config):
        """GET /api/v1/piles/configs/{pile_type}/ should return config."""
        response = api_client.get("/api/v1/piles/configs/TYPE_II/")
        assert response.status_code == status.HTTP_200_OK
        assert response.data["pile_type"] == "TYPE_II"
        assert len(response.data["main_bar_sections"]) == 2


class TestValidation:
    """Tests for input validation."""

    def test_duplicate_pile_no(self, api_client, project, pile_type_ii, type_ii_config):
        """Creating duplicate pile_no in same project should fail."""
        data = {
            "pile_no": pile_type_ii.pile_no,  # Duplicate
            "pile_type": "TYPE_II",
            "project": project.id,
            "diameter_mm": 500,
            "design_length_m": 20.0,
            "actual_length_m": 21.0,
        }
        response = api_client.post("/api/v1/piles/", data, format="json")
        assert response.status_code == status.HTTP_400_BAD_REQUEST


class TestAuthorization:
    """Tests for project-level authorization."""

    def test_engineer_cannot_list_unassigned_project(self, api_client, db):
        """Engineers should not see projects without membership."""
        Project.objects.create(name="Unassigned Project", status="ACTIVE")

        response = api_client.get("/api/v1/projects/")

        names = [item["name"] for item in response.data["results"]]
        assert "Unassigned Project" not in names

    def test_viewer_cannot_update_assigned_project(self, db, project):
        """Viewers should have read-only access to assigned projects."""
        user = get_user_model().objects.create_user(
            username="viewer-user",
            password="test-password",
        )
        viewer_group, _ = Group.objects.get_or_create(name="viewer")
        user.groups.add(viewer_group)
        ProjectMembership.objects.create(
            project=project,
            user=user,
            role=ProjectMembership.ROLE_VIEWER,
        )
        client = APIClient()
        client.force_authenticate(user=user)

        response = client.patch(
            f"/api/v1/projects/{project.id}/",
            {"name": "Viewer Update"},
            format="json",
        )

        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_engineer_cannot_create_pile_for_unassigned_project(
        self,
        api_client,
        db,
        type_ii_config,
    ):
        """Engineers should not write piles into projects they do not belong to."""
        unassigned_project = Project.objects.create(
            name="Unassigned Project",
            status="ACTIVE",
        )
        data = {
            "pile_no": "P-AUTH-001",
            "pile_type": "TYPE_II",
            "project": unassigned_project.id,
            "diameter_mm": 500,
            "design_length_m": 20.0,
            "actual_length_m": 21.0,
        }

        response = api_client.post("/api/v1/piles/", data, format="json")

        assert response.status_code == status.HTTP_400_BAD_REQUEST

    def test_invalid_bar_size(self):
        """Invalid bar size should raise ValueError."""
        from apps.piles.calculations import get_kg_per_m

        with pytest.raises(ValueError):
            get_kg_per_m(15)

    def test_invalid_pile_type(self, api_client, project):
        """Creating pile with invalid type should fail gracefully."""
        data = {
            "pile_no": "P-BAD",
            "pile_type": "INVALID_TYPE",
            "project": project.id,
            "diameter_mm": 500,
            "design_length_m": 20.0,
            "actual_length_m": 21.0,
        }
        response = api_client.post("/api/v1/piles/", data, format="json")
        assert response.status_code == status.HTTP_400_BAD_REQUEST
